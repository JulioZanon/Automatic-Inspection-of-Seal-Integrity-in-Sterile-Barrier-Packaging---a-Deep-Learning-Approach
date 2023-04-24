import numpy as np
import pandas as pd
from openpyxl import load_workbook
import os
from datetime import datetime

# todo: Summary Report: add as you go or save in case of error so that we do not lose all data if large train fails

def log_search_to_excel(traning_acc, validation_acc,traning_loss, validation_loss, trainable_layers=np.array([]),learning_rate_log=np.array([]),
                     path_from_project='logs\\xls', file_name_no_extension = 'performance_log',
                     search_info_l1 = '', search_info_l2 = '', worksheet_name = ''):
    #datetime.now().strftime('%Y%m%d-%H%M%S')
    #check that folder exists and create if not
    path = os.path.join(os.getcwd(), path_from_project)
    os.makedirs(path, exist_ok=True)
    file_name = file_name_no_extension + '.xlsx'
    file_path = os.path.join(path, file_name)
    #create numpy arrays from python lists
    traning_acc_np = np.array(traning_acc)
    validation_acc_np = np.array(validation_acc)
    traning_loss_np = np.array(traning_loss)
    validation_loss_np = np.array(validation_loss)
    if trainable_layers.shape[0] == 0 or learning_rate_log.shape[0] == 0:
        log_array = np.array(np.vstack((np.arange(traning_acc_np.shape[0]),
                                        traning_acc_np, validation_acc_np,
                                        traning_loss_np, validation_loss_np, )).T)
        log_header = ['Epochs','Training_ACC','VAL_ACC','Training_Loss','VAL_Loss']
    else:
        log_array = np.array(np.vstack((np.arange(traning_acc_np.shape[0]),
                                        traning_acc_np, validation_acc_np,
                                        traning_loss_np, validation_loss_np,
                                        trainable_layers,learning_rate_log)).T)
        log_header = ['Epochs', 'Training_ACC', 'VAL_ACC','Training_Loss','VAL_Loss', 'num_layers', 'lr']
    log_array_pd = pd.DataFrame(log_array)
    search_info = 'earch time stamp: [' +datetime.now().strftime("%d/%m/%Y, %H:%M:%S") + ']'
    search_info_pd = pd.DataFrame([search_info, search_info_l1, '', search_info_l2])
    search_num = 1

    if os.path.exists(file_path):
        # append worksheet with experiment search to existing eperiment workbook
        book = load_workbook(file_path)
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            search_num = len(writer.sheets) + 1
            if len(worksheet_name) == 0:
                worksheet_name = 'SEARCH-' + str(search_num)
            log_array_pd.to_excel(writer, worksheet_name, index=False, header=log_header)
            search_info_pd.to_excel(writer, worksheet_name, startrow=1, startcol=(log_array.shape[1] + 2),
                                            header=False, index=False)
            #book.worksheets[-1].cell(row=2, column=(log_array.shape[1]+2), value='Check')
    else:
        if len(worksheet_name) == 0:
            worksheet_name = 'SEARCH-' + str(search_num)
        #create new workbook with experiment SEARCH
        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
            log_array_pd.to_excel(writer,worksheet_name, index=False, header=log_header)
            search_info_pd.to_excel(writer, worksheet_name, startrow=1, startcol=(log_array.shape[1]+2),
                                            header=False, index=False)
    #with Pandas 1.3.4 save and close statements do not work as the file gets saved and close wtih the writing operation.
    #writer.save()
    #writer.close()


def log_summary_to_excel(name = 'Summary', sumary_matrix = [], header = [],
                     path_from_project='logs\\xls', file_name_no_extension = 'performance_log',
                     search_info_l1 = '', search_info_l2 = '',
                     search_info_l3 = 'Search information line 3'):

    #check that folder exists and create if not
    path = os.path.join(os.getcwd(), path_from_project)
    os.makedirs(path, exist_ok=True)
    file_name = file_name_no_extension + '.xlsx'
    file_path = os.path.join(path, file_name)
    #create numpy arrays from python lists
    log_array_pd = pd.DataFrame(sumary_matrix)
    search_info = 'Summary time stamp: [' +datetime.now().strftime("%d/%m/%Y, %H:%M:%S") + ']'
    search_info_pd = pd.DataFrame([search_info, search_info_l1, '', search_info_l2])

    if len(header) != sumary_matrix.shape[1]:
        print('log summary header incorrect ')
        print(str(len(header)))
        print(str(sumary_matrix.shape[1]))
    else:
        if os.path.exists(file_path):
            # append worksheet with experiment search to existing eperiment workbook
            book = load_workbook(file_path)
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                search_num = len(writer.sheets) + 1
                worksheet_name = name
                log_array_pd.to_excel(writer, worksheet_name, index=False, header=header)
                search_info_pd.to_excel(writer, worksheet_name, startrow=1, startcol=(sumary_matrix.shape[1] + 2),
                                                header=False, index=False)
                #book.worksheets[-1].cell(row=2, column=(log_array.shape[1]+2), value='Check')

                # with Pandas 1.3.4 save and close statements do not work as the file gets saved and close wtih the writing operation.
                #writer.save()
                #writer.close()
        else:
            worksheet_name = name
            #create new workbook with experiment SEARCH
            with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                log_array_pd.to_excel(writer,worksheet_name, index=False, header=header)
                search_info_pd.to_excel(writer, worksheet_name, startrow=1, startcol=(sumary_matrix.shape[1]+2),
                                                header=False, index=False)
        # with Pandas 1.3.4 save and close statements do not work as the file gets saved and close wtih the writing operation.
        #writer.save()
        #writer.close()

def log_test_results_to_excel(evaluation,merit, class_predictions, path_from_project='logs\\xls', file_name_no_extension = 'performance_log'):
    #check that folder exists and create if not
    path = os.path.join(os.getcwd(), path_from_project)
    os.makedirs(path, exist_ok=True)
    file_name = file_name_no_extension + '.xlsx'
    file_path = os.path.join(path, file_name)

    log_array_pd = pd.DataFrame(class_predictions)
    evaluation_info = 'evaluation time stamp: [' +datetime.now().strftime("%d/%m/%Y, %H:%M:%S") + ']'
    evaluation_info_l1 = 'Accuracy: [' + str(evaluation[1]) + ']. Loss: [' + str(evaluation[0]) + ']'
    evaluation_info_l2 = 'Min Score: [' + str(merit[0, 0] * 100) +  ']. Median Score: [' + str(merit[0, 1]* 100) + ']. Std Score: [' + str(merit[0, 2] * 100)+ ']'

    evaluation_info_pd = pd.DataFrame([evaluation_info, evaluation_info_l1, '', evaluation_info_l2])
    class_num = 0

    if os.path.exists(file_path):
        # append worksheet with experiment search to existing eperiment workbook
        book = load_workbook(file_path)
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            class_num = len(writer.sheets)
            worksheet_name = 'CLASS-' + str(class_num)
            log_array_pd.to_excel(writer, worksheet_name, index=False, header = ['img idx','true class score','false class score'])
            evaluation_info_pd.to_excel(writer, worksheet_name, startrow=1, startcol=(class_predictions.shape[1] + 2),
                                            header=False, index=False)
            #book.worksheets[-1].cell(row=2, column=(log_array.shape[1]+2), value='Check')
    else:
        worksheet_name = 'CLASS-' + str(class_num)
        #create new workbook with experiment SEARCH
        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
            log_array_pd.to_excel(writer,worksheet_name, index=False, header = ['img idx','true class score','false class score'])
            evaluation_info_pd.to_excel(writer, worksheet_name, startrow=1, startcol=(class_predictions.shape[1]+2),
                                            header=False, index=False)
    writer.save()
    writer.close()


def log_dictionary_to_excel(name = 'Summary', dictionary = dict(), path_from_project='logs\\xls', file_name_no_extension = 'performance_log',
                     info_l1 = '', info_l2 = ''):

    #check that folder exists and create if not
    path = os.path.join(os.getcwd(), path_from_project)
    os.makedirs(path, exist_ok=True)
    file_name = file_name_no_extension + '.xlsx'
    file_path = os.path.join(path, file_name)


    #create numpy arrays from python lists
    log_array_pd = pd.DataFrame.from_dict(dictionary, orient='index').transpose()
    search_info = 'time stamp: [' +datetime.now().strftime("%d/%m/%Y, %H:%M:%S") + ']'
    search_info_pd = pd.DataFrame([search_info, info_l1, '', info_l2])


    if os.path.exists(path):
        worksheet_name = name
        #create new workbook with experiment SEARCH
        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
            log_array_pd.to_excel(writer,worksheet_name, index=False, header=False)
            search_info_pd.to_excel(writer, worksheet_name, startrow=1, startcol=50,
                                            header=False, index=False)
        writer.save()
        writer.close()
    else: print('Report folder does not exist')


def display_dict(dictionary):
    for key, value in dictionary.items():
        if isinstance(value, dict):
            print(f'\n{key}:')
            display_dict(value)
        else:
            print(f'    >{key}: {value}')


def convert_cls_str_to_bool(np_class_array):
    for i in range(np_class_array.shape[0]):
        np_class_array[i] = str.replace(np_class_array[i], '[ ', '')
        np_class_array[i] = str.replace(np_class_array[i], '[', '')
        np_class_array[i] = str.replace(np_class_array[i], '] ', '')
        np_class_array[i] = str.replace(np_class_array[i], ']', '')
        np_class_array[i] = str.replace(np_class_array[i], 'True', '1')
        np_class_array[i] = str.replace(np_class_array[i], 'False', '0')
        if i == 0:
            class_array = np.asarray(np.fromstring(np_class_array[i], dtype=bool, sep=' '))
            class_array = np.expand_dims(class_array, axis=0)
        else:
            class_array = np.append(class_array,
                                    np.expand_dims(np.asarray(np.fromstring(np_class_array[i], dtype=bool, sep=' ')), axis=0),
                                    axis=0)
    return class_array
